"""Direct translation of the NHLBI extraction table from LaTeX to xlsx, for visual review.

HOW TO RUN
    python scripts/08_tex_to_xlsx.py

WHAT IT DOES
    Reads the newest Ground Truth Raw/crt_review_table_NNN.tex with the same parser
    07_build_ground_truth.py uses (imported, not reimplemented, so the two cannot drift
    apart) and writes it straight to xlsx: all 159 entries, all 22 columns, in tex
    order, one row per entry. No normalization, no join to paper_id, no dropped rows --
    this is the source table made readable, not the merged ground truth. For that, see
    data/ground_truth.csv.

    Cited-but-unreviewed entries (23 of them -- a citation with every field blank) are
    left blank here too, and flagged in the `reviewed` column so they read as "not done"
    rather than "reviewed, nothing to report".

OUTPUT
    Ground Truth Raw/crt_review_table_112.xlsx
"""

import importlib.util
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / "Ground Truth Raw"

# Import read_nhlbi/newest_tex/TEX_COLUMNS from 07 by file path -- its filename starts
# with a digit, so it cannot be `import`ed by name.
spec = importlib.util.spec_from_file_location("build_ground_truth", Path(__file__).parent / "07_build_ground_truth.py")
build_ground_truth = importlib.util.module_from_spec(spec)
sys.modules["build_ground_truth"] = build_ground_truth
spec.loader.exec_module(build_ground_truth)

HEADERS = [
    "entry", "citation", "exclude_reason", "n_trt", "n_levels", "comment_levels",
    "n_outer", "n_2nd", "n_3rd", "n_4th", "unit_rand", "ind_samp_unit",
    "restricted_rand", "icc", "n_long", "stepped_wedge", "data_done", "data_should",
    "data_correct", "data_comment", "power_done", "power_should", "power_correct",
    "reviewed", "note",
]


def main() -> None:
    tex_path = build_ground_truth.newest_tex()
    rows = build_ground_truth.read_nhlbi(tex_path)
    print(f"read {len(rows)} entries from {tex_path.name}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "crt_review_table"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "C2"

    unreviewed_fill = PatternFill("solid", fgColor="F2F2F2")
    excluded_fill = PatternFill("solid", fgColor="FCE4D6")

    # read_nhlbi() stores five of these under a "_raw" suffix internally (it keeps
    # both a raw and a normalized copy for the merged CSV); this table wants the
    # source value regardless of which key it lives under.
    field_keys = {
        "exclude_reason": "exclusion_reason_raw",
        "restricted_rand": "restricted_rand_raw",
        "stepped_wedge": "stepped_wedge_raw",
        "data_correct": "data_correct_raw",
        "power_correct": "power_correct_raw",
    }

    for row in rows:
        values = [row.get(field_keys.get(c, c), "") for c in HEADERS[1:-2]]
        reviewed = "no" if not any(v for v in values[1:]) else "yes"
        excel_row = [row["source_row"], row["citation_raw"]] + values[1:] + [reviewed, row.get("source_note", "")]
        ws.append(excel_row)

        r = ws.max_row
        if reviewed == "no":
            for cell in ws[r]:
                cell.fill = unreviewed_fill
        elif values[1]:  # exclude_reason filled
            for cell in ws[r]:
                cell.fill = excluded_fill

    widths = {"citation": 26, "comment_levels": 30, "restricted_rand": 24, "icc": 18,
              "data_done": 34, "data_should": 30, "data_comment": 30,
              "power_done": 34, "power_should": 30, "note": 22}
    for i, name in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 12)

    ws.auto_filter.ref = ws.dimensions

    out_path = RAW / tex_path.with_suffix(".xlsx").name
    wb.save(out_path)
    print(f"wrote {out_path.relative_to(ROOT)}  ({ws.max_row - 1} rows)")
    print("shaded rows: grey = cited but not yet reviewed, orange = excluded")


if __name__ == "__main__":
    main()
